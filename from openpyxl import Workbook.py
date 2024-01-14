from openpyxl import Workbook
import array as arr

# 创建一个 workbook
wb = Workbook()
# 获取被激活的 worksheet
ws = wb.active
# 设置单元格内容
# ws['A1'] = 42
# 设置一行内容
# ws.append([1, 2, 3])
# python 数据类型可以被自动转换
# import datetime
# ws['A3'] = datetime.datetime.now()

# 建立表格内容的表头
ws['A1'] = 'id';ws['B1'] = 'name';ws['C1'] = 'gender';ws['D1'] = 'age';

id = (1,2,3,4)
name = ('Alice','Bob','Caty','Dota')
gender = ('girl','boy','girl','boy')
age = (18,20,19,21)

i = 0
cella = ws['A2:A5']
# print(cella)

# 填入id
for row in cella:
    row = name[i]
    i = i + 1
    print(i,row)
# for i in r:
#     print(i,id[i],name[i],gender[i],age[i])

# 保存 Excel 文件
wb.save("sample.xlsx")